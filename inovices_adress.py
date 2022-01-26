import os
from tkinter import Tk, filedialog
from tkinter import *
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.layout import LAParams
from pdfminer.converter import TextConverter
from io import StringIO
from pdfminer.pdfpage import PDFPage
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

root = Tk()
root.geometry("400x400")


def open_dir():
    global dir
    dir = filedialog.askdirectory(initialdir="/home/krzysiek/Dokumenty")
    return dir


my_button = Button(root, text="Wybierz folder", command=open_dir).pack()
root.mainloop()
path = dir
filelist = []
for root, dirs, files in os.walk(path):
    for file in files:
        filelist.append(file)
    inovice_list_dir = []
    inovice_list = []
    for files in range(len(filelist)):
        if (filelist[files][:1]) == "F":
            inovice_list_dir.append(os.path.join(root, filelist[files]))
            inovice_list.append(filelist[files])
    inovice_list.sort()
    inovice_list_dir.sort()




def get_pdf_file_content(path_to_pdf):
    resource_manager = PDFResourceManager(caching=True)
    out_text = StringIO()
    codec = 'utf-8'
    laParams = LAParams()
    text_converter = TextConverter(resource_manager, out_text, laparams=laParams)
    fp = open(path_to_pdf, 'rb')
    interpreter = PDFPageInterpreter(resource_manager, text_converter)
    for page in PDFPage.get_pages(fp, pagenos=set(), maxpages=0, password="", caching=True, check_extractable=True):
        interpreter.process_page(page)
    text = out_text.getvalue()
    fp.close()
    text_converter.close()
    out_text.close()
    return text


inovice_nr_list=[]
inovice_adress_list=[]



for i in range (len(inovice_list_dir)):
    tekst = get_pdf_file_content(inovice_list_dir[i])
    inovice_nr = tekst.split("/2021")[0]
    inovice_nr = inovice_nr.split("FV")[1]
    inovice_nr_list.append(inovice_nr)
    inovice_adress = tekst.split("Data wystawienia")[0]
    inovice_adress = inovice_adress.split("\n\n")
    inovice_adress = inovice_adress[(len(inovice_adress)-2)]
    inovice_adress = inovice_adress.replace("SPÓŁKA Z\nOGRANICZONĄ ODPOWIEDZIALNOŚCIĄ","sp. z o.o.")
    inovice_adress = inovice_adress.replace("SPÓŁKA\nZ OGRANICZONĄ ODPOWIEDZIALNOŚCIĄ", "sp. z o.o.")
    inovice_adress = inovice_adress.replace("SPÓŁKA Z OGRANICZONĄ\nODPOWIEDZIALNOŚCIĄ", "sp. z o.o.")
    inovice_adress = inovice_adress.replace("SPÓŁKA KOMANDYTOWA", "sp. k.")
    inovice_adress = inovice_adress.replace("SPÓŁKA\n KOMANDYTOWA", "sp. k.")
    inovice_adress = inovice_adress.replace("SPÓŁKA CYWILNA", "s.c.")
    inovice_adress_list.append(inovice_adress)



inovice_adress_name = []
inovice_adress_street = []
inovice_adress_postcode = []
inovice_adress_city = []
inovice_adress_country = []

for i in range(len(inovice_adress_list)):
    inovice_adress_list[i] = inovice_adress_list[i].split("\n")
    inovice_adress_list[i].pop(len(inovice_adress_list[i]) - 1)

for i in range(len(inovice_adress_list)):

    if inovice_adress_list[i][(len(inovice_adress_list[i]) - 1)][0:1].isdigit() and \
            inovice_adress_list[i][(len(inovice_adress_list[i]) - 1)][2] == "-" and inovice_adress_list[i][(
            len(inovice_adress_list[i]) - 1)][3:6].isdigit():
            inovice_adress_country.append("Poland")
            inovice_adress_postcode.append(inovice_adress_list[i][len(inovice_adress_list[i])-1][0:6])
            inovice_adress_city.append(inovice_adress_list[i] [len(inovice_adress_list[i]) - 1] \
                                           [6:(len(inovice_adress_list[i][len(inovice_adress_list[i])-1]))])
            inovice_adress_street.append(inovice_adress_list[i] [len(inovice_adress_list[i])-2])
            inovice_adress_name.append(inovice_adress_list[i][: len(inovice_adress_list[i]) - 2])



    else:
        inovice_adress_country.append("foreign contry")
        inovice_adress_postcode.append("foreign postcode")
        inovice_adress_city.append("foreign city")
        inovice_adress_street.append("foreing steet")
        inovice_adress_name.append("foreing name")


for i in range (len(inovice_adress_name)):
	inovice_adress_name[i] = ''.join(map(str, inovice_adress_name[i]))


print (inovice_adress_name)
print (inovice_adress_street)
print (inovice_adress_postcode)
print (inovice_nr_list)
print(inovice_adress_name)

for i in range (len(inovice_nr_list)):
    print(inovice_nr_list[i])

wb = openpyxl.load_workbook("/home/krzysiek/Dokumenty/rk369669552fr/adresowanie kopert.xlsx")
ws = wb['Arkusz1']

for i in range (len(inovice_list)):
    ws.cell(column=1, row=i + 2).value = int(inovice_nr_list[i])
    ws.cell(column=2, row=i + 2).value = str(inovice_adress_name[i])
    ws.cell(column=3, row=i + 2).value = str(inovice_adress_street[i])
    ws.cell(column=4, row=i + 2).value = str(inovice_adress_postcode[i])
    ws.cell(column=5, row=i + 2).value = str(inovice_adress_city[i])
    if inovice_adress_country[i] == "Poland":
        ws.cell(column=6, row=i + 2).value = ""
    else:
        ws.cell(column=6, row=i + 2).value = str(inovice_adress_country[i])




wb.save("/home/krzysiek/Dokumenty/rk369669552fr/adresowanie kopert.xlsx")
